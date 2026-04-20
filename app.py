import io
import json
import re
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="Kurier Analysis", layout="centered")

st.title("Kurier Analysis")


RECENT_SOURCES_PATH = Path(__file__).with_name("recent_sources.json")
MAX_RECENT_SOURCES = 3


month_map = {
    "jan": "jan", "january": "jan",
    "feb": "feb", "february": "feb", "febrauray": "feb", "febraury": "feb",
    "mar": "mar", "march": "mar",
    "apr": "apr", "april": "apr",
    "may": "may",
    "jun": "jun", "june": "jun",
    "jul": "jul", "july": "jul",
    "aug": "aug", "august": "aug",
    "sep": "sep", "september": "sep",
    "oct": "oct", "october": "oct",
    "nov": "nov", "november": "nov",
    "dec": "dec", "december": "dec",
}


def is_valid_year(y):
    y = int(y)
    if 2020 <= y <= 2030:
        return True
    if 20 <= y <= 30:
        return True
    return False


def is_valid_month_year_sheet(name):
    name = name.strip().lower()

    if not re.fullmatch(r"[a-z0-9\s]+", name):
        return False

    parts = name.split()

    if len(parts) != 2:
        return False

    part1, part2 = parts

    if part1 in month_map and part2.isdigit() and is_valid_year(part2):
        return True

    if part2 in month_map and part1.isdigit() and is_valid_year(part1):
        return True

    return False


def get_excel_source(excel_url, uploaded_file):
    if uploaded_file is not None:
        return io.BytesIO(uploaded_file.getvalue())

    if not excel_url:
        raise ValueError("Please paste a Google Sheet / Excel URL or upload an Excel file.")

    excel_source = excel_url.strip()
    if "docs.google.com/spreadsheets" in excel_source and "/edit" in excel_source:
        return excel_source.split("/edit")[0] + "/export?format=xlsx"

    return excel_source


def get_workbook_title_from_excel_source(excel_source):
    try:
        if isinstance(excel_source, io.BytesIO):
            workbook = load_workbook(io.BytesIO(excel_source.getvalue()), read_only=True)
        else:
            workbook = load_workbook(excel_source, read_only=True)

        title = getattr(workbook.properties, "title", None)
        return title.strip() if isinstance(title, str) and title.strip() else None
    except Exception:
        return None


def get_google_sheet_title_from_url(excel_url):
    if not excel_url or "docs.google.com/spreadsheets" not in excel_url:
        return None

    page_url = excel_url.strip()
    if "/export" in page_url:
        page_url = page_url.split("/export")[0] + "/edit"

    try:
        request = Request(page_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=10) as response:
            html = response.read().decode("utf-8", errors="ignore")

        match = re.search(r"<title>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
        if not match:
            return None

        title = re.sub(r"\s+", " ", match.group(1)).strip()
        return title.removesuffix(" - Google Sheets").strip()
    except Exception:
        return None


def load_all_sheets(excel_url, uploaded_file):
    excel_source = get_excel_source(excel_url, uploaded_file)
    all_sheets = pd.read_excel(excel_source, sheet_name=None, engine="openpyxl")
    workbook_title = get_workbook_title_from_excel_source(excel_source)
    return all_sheets, workbook_title


def get_monthly_sheets_filtered(all_sheets):
    return {
        name: df for name, df in all_sheets.items()
        if is_valid_month_year_sheet(name)
    }


def get_spreadsheet_display_name(excel_url, uploaded_file, workbook_title=None):
    if workbook_title:
        return workbook_title

    google_sheet_title = get_google_sheet_title_from_url(excel_url)
    if google_sheet_title:
        return google_sheet_title

    if uploaded_file is not None:
        return Path(uploaded_file.name).stem

    if not excel_url:
        return None

    source = excel_url.strip()
    parsed_url = urlparse(source)
    file_name = unquote(parsed_url.path.rstrip("/").split("/")[-1])
    return Path(file_name).stem if file_name else source


def load_recent_sources():
    if not RECENT_SOURCES_PATH.exists():
        return []

    try:
        saved_sources = json.loads(RECENT_SOURCES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    if not isinstance(saved_sources, list):
        return []

    recent_sources = []
    for entry in saved_sources:
        if not isinstance(entry, dict):
            continue

        url = str(entry.get("url", "")).strip()
        name = str(entry.get("name", "")).strip()

        if not url:
            continue

        recent_sources.append({
            "name": name or get_spreadsheet_display_name(url, None),
            "url": url,
        })

    return recent_sources[:MAX_RECENT_SOURCES]


def save_recent_source(name, url):
    cleaned_url = (url or "").strip()
    if not cleaned_url:
        return

    cleaned_name = (name or "").strip() or get_spreadsheet_display_name(cleaned_url, None)
    recent_sources = [
        entry for entry in load_recent_sources()
        if entry.get("url", "").strip() != cleaned_url
    ]

    recent_sources.insert(0, {
        "name": cleaned_name,
        "url": cleaned_url,
    })

    RECENT_SOURCES_PATH.write_text(
        json.dumps(recent_sources[:MAX_RECENT_SOURCES], indent=2),
        encoding="utf-8",
    )


def add_serial_number(df):
    display_df = df.reset_index(drop=True).copy()
    display_df.insert(0, "S. No.", [str(i) for i in range(1, len(display_df) + 1)])
    return display_df


def prepare_display_table(df, left_align_packet_count=False, left_align_payment_columns=False):
    display_df = add_serial_number(df)
    if left_align_packet_count and "Packet Count" in display_df.columns:
        display_df["Packet Count"] = display_df["Packet Count"].astype(str)
    if left_align_payment_columns:
        for col in ["CASH AMOUNT", "UPI AMOUNT", "CREDIT AMOUNT", "CREDIT COUNT", "TRANSACTION COUNT"]:
            if col in display_df.columns:
                display_df[col] = display_df[col].astype(str)
    return display_df


def get_table_column_config(left_align_packet_count=False, left_align_payment_columns=False):
    column_config = {
        "S. No.": st.column_config.TextColumn(
            "S. No.",
            width="small",
        )
    }
    if left_align_packet_count:
        column_config["Packet Count"] = st.column_config.TextColumn("Packet Count")
    if left_align_payment_columns:
        column_config["CASH AMOUNT"] = st.column_config.TextColumn("CASH AMOUNT")
        column_config["UPI AMOUNT"] = st.column_config.TextColumn("UPI AMOUNT")
        column_config["CREDIT AMOUNT"] = st.column_config.TextColumn("CREDIT AMOUNT")
        column_config["CREDIT COUNT"] = st.column_config.TextColumn("CREDIT COUNT")
        column_config["TRANSACTION COUNT"] = st.column_config.TextColumn("TRANSACTION COUNT")
    return column_config


def get_date_wise_packet_count(sheet_name, monthly_sheets_filtered):
    df = monthly_sheets_filtered[sheet_name].copy()

    df.columns = df.columns.astype(str).str.strip().str.upper()
    df = df.loc[:, ~df.columns.duplicated()]

    if "DATE" in df.columns:
        date_col = "DATE"
    elif "AHU" in df.columns:
        date_col = "AHU"
    else:
        date_col = df.columns[0]

    df["DATE"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    df = df[df["DATE"].notna()]

    if "AWB NO." not in df.columns:
        return pd.DataFrame()

    result = (
        df.groupby("DATE")["AWB NO."]
        .count()
        .reset_index(name="Packet Count")
        .sort_values("DATE", ascending=True)
    )

    result["DATE"] = result["DATE"].dt.date
    return result


def get_packets_booked_per_sender(sheet_name, monthly_sheets_filtered):
    df = monthly_sheets_filtered[sheet_name].copy()

    df.columns = df.columns.astype(str).str.strip().str.upper()
    df = df.loc[:, ~df.columns.duplicated()]

    if "AWB NO." not in df.columns:
        return pd.DataFrame()

    if "SENDER NAME" in df.columns:
        sender_col = "SENDER NAME"
    elif "MODE" in df.columns:
        sender_col = "MODE"
    elif "BILLING DETAILS" in df.columns:
        sender_col = "BILLING DETAILS"
    else:
        return pd.DataFrame()

    df[sender_col] = df[sender_col].astype(str).str.strip()

    result = (
        df.groupby(sender_col)["AWB NO."]
        .count()
        .reset_index(name="Packet Count")
        .sort_values("Packet Count", ascending=False)
    )

    return result.rename(columns={sender_col: "SENDER NAME"})


def get_sender_wise_packets_for_each_date(sheet_name, monthly_sheets_filtered):
    df = monthly_sheets_filtered[sheet_name].copy()

    df.columns = df.columns.astype(str).str.strip().str.upper()
    df = df.loc[:, ~df.columns.duplicated()]

    if "DATE" in df.columns:
        date_col = "DATE"
    elif "AHU" in df.columns:
        date_col = "AHU"
    else:
        date_col = df.columns[0]

    df["DATE"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    df = df[df["DATE"].notna()]

    required_cols = ["DATE", "SENDER NAME", "AWB NO."]
    if not all(col in df.columns for col in required_cols):
        return pd.DataFrame()

    result = (
        df.groupby(["DATE", "SENDER NAME"])["AWB NO."]
        .count()
        .reset_index(name="Packet Count")
        .sort_values(["DATE", "Packet Count", "SENDER NAME"], ascending=[True, False, True])
    )

    result["DATE"] = result["DATE"].dt.date
    return result


def get_packets_booked_per_mode(sheet_name, monthly_sheets_filtered):
    df = monthly_sheets_filtered[sheet_name].copy()

    df.columns = df.columns.astype(str).str.strip().str.upper()
    df = df.loc[:, ~df.columns.duplicated()]

    if "AWB NO." not in df.columns or "MODE" not in df.columns:
        return pd.DataFrame()

    df["MODE"] = df["MODE"].astype(str).str.upper().str.strip()

    return (
        df.groupby("MODE")["AWB NO."]
        .count()
        .reset_index(name="Packet Count")
        .sort_values("Packet Count", ascending=False)
    )


def get_payment_base_df(sheet_name, monthly_sheets_filtered):
    df = monthly_sheets_filtered[sheet_name].copy()

    df.columns = df.columns.astype(str).str.strip().str.upper()
    df = df.loc[:, ~df.columns.duplicated()]

    if "DATE" in df.columns:
        date_col = "DATE"
    elif "AHU" in df.columns:
        date_col = "AHU"
    else:
        date_col = df.columns[0]

    required_cols = ["CREDIT OR CASH", "AMOUNT"]
    if not all(col in df.columns for col in required_cols):
        return None

    df["DATE"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    df = df[df["DATE"].notna()]
    df["DATE"] = df["DATE"].dt.normalize()
    df["CREDIT OR CASH"] = df["CREDIT OR CASH"].astype(str).str.upper().str.strip()
    df["AMOUNT"] = df["AMOUNT"].astype(str).str.strip()
    df["AMOUNT_NUM"] = pd.to_numeric(df["AMOUNT"], errors="coerce")
    if "SENDER NAME" in df.columns:
        df["SENDER NAME"] = df["SENDER NAME"].astype(str).str.strip()

    return df


def get_payment_received_per_month(sheet_name, monthly_sheets_filtered):
    df = get_payment_base_df(sheet_name, monthly_sheets_filtered)

    if df is None:
        return None

    if df.empty:
        return pd.DataFrame(
            columns=[
                "DATE",
                "CASH AMOUNT",
                "UPI AMOUNT",
                "CREDIT AMOUNT",
                "CREDIT COUNT",
                "TRANSACTION COUNT",
            ]
        )

    all_dates = pd.DataFrame({
        "DATE": sorted(df["DATE"].unique())
    })

    cash_amount = (
        df[df["CREDIT OR CASH"] == "CASH"]
        .groupby("DATE")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="CASH AMOUNT")
    )

    upi_amount = (
        df[df["CREDIT OR CASH"] == "UPI"]
        .groupby("DATE")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="UPI AMOUNT")
    )

    credit_amount = (
        df[df["CREDIT OR CASH"] == "CREDIT"]
        .groupby("DATE")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="CREDIT AMOUNT")
    )

    credit_count = (
        df[
            (df["CREDIT OR CASH"] == "CREDIT") &
            (df["AMOUNT"].str.lower() == "monthly")
        ]
        .groupby("DATE")["AMOUNT"]
        .count()
        .reset_index(name="CREDIT COUNT")
    )

    transaction_count = (
        df[df["CREDIT OR CASH"].isin(["CASH", "UPI", "CREDIT"])]
        .groupby("DATE")["CREDIT OR CASH"]
        .count()
        .reset_index(name="TRANSACTION COUNT")
    )

    result = (
        all_dates
        .merge(cash_amount, on="DATE", how="left")
        .merge(upi_amount, on="DATE", how="left")
        .merge(credit_amount, on="DATE", how="left")
        .merge(credit_count, on="DATE", how="left")
        .merge(transaction_count, on="DATE", how="left")
        .fillna(0)
        .sort_values("DATE", ascending=True)
    )

    result["DATE"] = result["DATE"].dt.date

    for col in ["CASH AMOUNT", "UPI AMOUNT", "CREDIT AMOUNT"]:
        result[col] = result[col].apply(
            lambda value: int(value) if float(value).is_integer() else round(float(value), 2)
        )

    result["CREDIT COUNT"] = result["CREDIT COUNT"].astype(int)
    result["TRANSACTION COUNT"] = result["TRANSACTION COUNT"].astype(int)

    return result


def get_sender_wise_payment_for_date(sheet_name, monthly_sheets_filtered, selected_date):
    df = get_payment_base_df(sheet_name, monthly_sheets_filtered)

    if df is None or "SENDER NAME" not in df.columns:
        return None

    df = df[
        (df["SENDER NAME"] != "") &
        (df["SENDER NAME"].str.lower() != "nan")
    ]

    if df.empty:
        return pd.DataFrame(
            columns=[
                "SENDER NAME",
                "CASH AMOUNT",
                "UPI AMOUNT",
                "CREDIT AMOUNT",
                "CREDIT COUNT",
                "TRANSACTION COUNT",
            ]
        )

    selected_date = pd.to_datetime(selected_date).normalize()
    df = df[df["DATE"] == selected_date]

    if df.empty:
        return pd.DataFrame(
            columns=[
                "SENDER NAME",
                "CASH AMOUNT",
                "UPI AMOUNT",
                "CREDIT AMOUNT",
                "CREDIT COUNT",
                "TRANSACTION COUNT",
            ]
        )

    all_senders = pd.DataFrame({
        "SENDER NAME": sorted(df["SENDER NAME"].unique())
    })

    cash_amount = (
        df[df["CREDIT OR CASH"] == "CASH"]
        .groupby("SENDER NAME")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="CASH AMOUNT")
    )

    upi_amount = (
        df[df["CREDIT OR CASH"] == "UPI"]
        .groupby("SENDER NAME")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="UPI AMOUNT")
    )

    credit_amount = (
        df[df["CREDIT OR CASH"] == "CREDIT"]
        .groupby("SENDER NAME")["AMOUNT_NUM"]
        .sum()
        .reset_index(name="CREDIT AMOUNT")
    )

    credit_count = (
        df[
            (df["CREDIT OR CASH"] == "CREDIT") &
            (df["AMOUNT"].str.lower() == "monthly")
        ]
        .groupby("SENDER NAME")["AMOUNT"]
        .count()
        .reset_index(name="CREDIT COUNT")
    )

    transaction_count = (
        df[df["CREDIT OR CASH"].isin(["CASH", "UPI", "CREDIT"])]
        .groupby("SENDER NAME")["CREDIT OR CASH"]
        .count()
        .reset_index(name="TRANSACTION COUNT")
    )

    result = (
        all_senders
        .merge(cash_amount, on="SENDER NAME", how="left")
        .merge(upi_amount, on="SENDER NAME", how="left")
        .merge(credit_amount, on="SENDER NAME", how="left")
        .merge(credit_count, on="SENDER NAME", how="left")
        .merge(transaction_count, on="SENDER NAME", how="left")
        .fillna(0)
        .sort_values("SENDER NAME", ascending=True)
    )

    for col in ["CASH AMOUNT", "UPI AMOUNT", "CREDIT AMOUNT"]:
        result[col] = result[col].apply(
            lambda value: int(value) if float(value).is_integer() else round(float(value), 2)
        )

    result["CREDIT COUNT"] = result["CREDIT COUNT"].astype(int)
    result["TRANSACTION COUNT"] = result["TRANSACTION COUNT"].astype(int)

    return result


def apply_recent_source_selection():
    selected_recent_url = st.session_state.recent_source_url
    if selected_recent_url:
        st.session_state.excel_url_input = selected_recent_url


if "monthly_sheets_filtered" not in st.session_state:
    st.session_state.monthly_sheets_filtered = None

if "confirmed_month" not in st.session_state:
    st.session_state.confirmed_month = None

if "confirmed_report" not in st.session_state:
    st.session_state.confirmed_report = None

if "spreadsheet_name" not in st.session_state:
    st.session_state.spreadsheet_name = None

if "excel_url_input" not in st.session_state:
    st.session_state.excel_url_input = ""

if "recent_source_url" not in st.session_state:
    st.session_state.recent_source_url = ""


recent_sources = load_recent_sources()
if recent_sources:
    recent_source_names = {
        entry["url"]: entry["name"] for entry in recent_sources
    }
    st.selectbox(
        "Saved Recent Sheets",
        options=[""] + [entry["url"] for entry in recent_sources],
        format_func=lambda url: "Select saved sheet" if not url else recent_source_names.get(url, url),
        key="recent_source_url",
        on_change=apply_recent_source_selection,
    )
    st.caption("The latest 3 working URLs are saved automatically.")


excel_url = st.text_input("Paste Google Sheet / Excel URL", key="excel_url_input")
uploaded_file = st.file_uploader("Or upload Excel file", type=["xlsx", "xls"])

if st.button("Analyze", key="source_analyze"):
    try:
        all_sheets, workbook_title = load_all_sheets(excel_url, uploaded_file)
        monthly_sheets_filtered = get_monthly_sheets_filtered(all_sheets)

        if not monthly_sheets_filtered:
            st.session_state.monthly_sheets_filtered = None
            st.session_state.confirmed_month = None
            st.session_state.confirmed_report = None
            st.session_state.spreadsheet_name = None
            st.error("No valid month sheets were found in the file.")
        else:
            st.session_state.monthly_sheets_filtered = monthly_sheets_filtered
            st.session_state.confirmed_month = None
            st.session_state.confirmed_report = None
            st.session_state.spreadsheet_name = get_spreadsheet_display_name(
                excel_url,
                uploaded_file,
                workbook_title=workbook_title,
            )
            if uploaded_file is None and excel_url.strip():
                save_recent_source(st.session_state.spreadsheet_name, excel_url)
                st.session_state.recent_source_url = excel_url.strip()
            st.success("File analyzed. Select a month below.")
    except Exception as exc:
        st.session_state.monthly_sheets_filtered = None
        st.session_state.confirmed_month = None
        st.session_state.confirmed_report = None
        st.session_state.spreadsheet_name = None
        st.error(f"Could not read the sheet: {exc}")


if st.session_state.monthly_sheets_filtered:
    if st.session_state.spreadsheet_name:
        st.markdown(
            f"<div style='font-size: 1.2rem; font-weight: 600; margin-bottom: 0.5rem;'>"
            f"Spreadsheet: {st.session_state.spreadsheet_name}"
            f"</div>",
            unsafe_allow_html=True,
        )

    month_options = list(st.session_state.monthly_sheets_filtered.keys())
    selected_month = st.selectbox("Select Month", month_options)

    if st.button("Analyze", key="month_analyze"):
        st.session_state.confirmed_month = selected_month
        st.session_state.confirmed_report = None


if st.session_state.confirmed_month:
    report_options = [
        "date wise packet count",
        "month-wise packets booked per sender",
        "month-wise packets booked per mode",
        "payment received per month",
    ]

    selected_report = st.selectbox("Select Report", report_options)

    if st.button("Show Result"):
        st.session_state.confirmed_report = selected_report


if st.session_state.confirmed_month and st.session_state.confirmed_report:
    month_name = st.session_state.confirmed_month
    report_name = st.session_state.confirmed_report
    monthly_sheets_filtered = st.session_state.monthly_sheets_filtered

    if report_name == "date wise packet count":
        st.subheader("Date Wise Packet Count")
        result = get_date_wise_packet_count(month_name, monthly_sheets_filtered)
        if result.empty:
            st.warning("Required columns were not found for this report.")
        else:
            st.metric("Total Packets Booked This Month", int(result["Packet Count"].sum()))
            sender_date_result = get_sender_wise_packets_for_each_date(month_name, monthly_sheets_filtered)
            display_result = prepare_display_table(result, left_align_packet_count=True)
            event = st.dataframe(
                display_result,
                use_container_width=True,
                hide_index=True,
                column_config=get_table_column_config(left_align_packet_count=True),
                on_select="rerun",
                selection_mode="single-row",
                key="date_wise_table",
            )

            if not sender_date_result.empty:
                selected_rows = event.selection.rows
                if selected_rows:
                    selected_row = result.iloc[selected_rows[0]]
                    selected_date = selected_row["DATE"]
                    day_result = sender_date_result[sender_date_result["DATE"] == selected_date]

                    st.write(f"Sender-wise packet count for {selected_date}")
                    st.dataframe(
                        prepare_display_table(day_result[["SENDER NAME", "Packet Count"]], left_align_packet_count=True),
                        use_container_width=True,
                        hide_index=True,
                        column_config=get_table_column_config(left_align_packet_count=True),
                    )
                else:
                    st.caption("Click a date row in the table to see sender-wise packet count for that day.")

    elif report_name == "month-wise packets booked per sender":
        st.subheader("Packets Booked Per Sender")
        result = get_packets_booked_per_sender(month_name, monthly_sheets_filtered)
        if result.empty:
            st.warning("Required columns were not found for this report.")
        else:
            st.metric("Total Packets Booked This Month", int(result["Packet Count"].sum()))
            st.dataframe(
                prepare_display_table(result, left_align_packet_count=True),
                use_container_width=True,
                hide_index=True,
                column_config=get_table_column_config(left_align_packet_count=True),
            )

    elif report_name == "month-wise packets booked per mode":
        st.subheader("Packets Booked Per Mode")
        result = get_packets_booked_per_mode(month_name, monthly_sheets_filtered)
        if result.empty:
            st.warning("Required columns were not found for this report.")
        else:
            st.dataframe(
                prepare_display_table(result, left_align_packet_count=True),
                use_container_width=True,
                hide_index=True,
                column_config=get_table_column_config(left_align_packet_count=True),
            )

    elif report_name == "payment received per month":
        st.subheader("Payment Received Per Month")
        result = get_payment_received_per_month(month_name, monthly_sheets_filtered)

        if result is None:
            st.warning("Required columns were not found for this report.")
        elif result.empty:
            st.info("No date-wise payment data found.")
        else:
            total_cash = result["CASH AMOUNT"].sum()
            total_upi = result["UPI AMOUNT"].sum()
            total_credit_amount = result["CREDIT AMOUNT"].sum()
            total_credit_count = int(result["CREDIT COUNT"].sum())
            total_transaction_count = int(result["TRANSACTION COUNT"].sum())

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Total Cash", total_cash)
            col2.metric("Total UPI", total_upi)
            col3.metric("Total Credit", total_credit_amount)
            col4.metric("Credit Count", total_credit_count)
            col5.metric("Transaction Count", total_transaction_count)

            event = st.dataframe(
                prepare_display_table(result, left_align_payment_columns=True),
                use_container_width=True,
                hide_index=True,
                column_config=get_table_column_config(left_align_payment_columns=True),
                on_select="rerun",
                selection_mode="single-row",
                key="payment_date_table",
            )

            selected_rows = event.selection.rows
            if selected_rows:
                selected_row = result.iloc[selected_rows[0]]
                selected_date = selected_row["DATE"]
                sender_result = get_sender_wise_payment_for_date(
                    month_name,
                    monthly_sheets_filtered,
                    selected_date,
                )

                if sender_result is None:
                    st.warning("SENDER NAME column was not found for the selected day's breakdown.")
                elif sender_result.empty:
                    st.info("No sender-wise payment data found for the selected date.")
                else:
                    st.write(f"Sender-wise payment details for {selected_date}")
                    st.dataframe(
                        prepare_display_table(sender_result, left_align_payment_columns=True),
                        use_container_width=True,
                        hide_index=True,
                        column_config=get_table_column_config(left_align_payment_columns=True),
                    )
            else:
                st.caption("Click a date row in the table to see sender-wise payment details for that day.")
